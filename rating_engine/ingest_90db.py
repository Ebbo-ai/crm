#!/usr/bin/env python3
"""
Simple Benefits - 90 Degree Benefits monthly feed ingest.

Turns the combined monthly feed (all groups, one row per group/plan/month) into
PPR payloads ready for generate_ppr.py. Handles the "update the prior PPR" step:
merges the new month into stored history, detects restatements of prior months,
and validates before anything is published.

Rate tables are NOT in the feed - they come from the CRM (set at renewal).

Usage:
    python3 ingest_90db.py feed_2026_05.csv rates.json store.json
        --> writes updated store.json, prints a reconciliation report

Design notes:
  * Store is append/merge by (group_id, plan_id, month). Never destructive.
  * A month whose figures move is recorded as a restatement with the prior
    values retained, so a report issued earlier stays reproducible.
  * Validation failures HOLD the row; they do not silently pass through.
"""
__version__ = "1.0"

import csv, json, sys, os, datetime as _dt
from collections import defaultdict

TIER_COLS = {
    "enroll_ee":         "ee",
    "enroll_ee_one":     "ee1",
    "enroll_ee_child":   "ec",
    "enroll_ee_spouse":  "es",
    "enroll_ee_family":  "fam",
}

def _f(v):
    if v is None: return 0.0
    s = str(v).strip().replace("$","").replace(",","")
    if s in ("", "-", "n/a", "N/A"): return 0.0
    neg = s.startswith("(") and s.endswith(")")
    if neg: s = s[1:-1]
    try: return -float(s) if neg else float(s)
    except ValueError: return 0.0

def _i(v):
    return int(round(_f(v)))


# ------------------------------------------------------------------ read feed
def read_feed(path):
    """Read the combined monthly feed. Accepts .csv or .xlsx."""
    rows = []
    if path.lower().endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        hdr = [str(h).strip().lower() if h is not None else "" for h in next(it)]
        for r in it:
            if r is None or all(c is None for c in r): continue
            rows.append({hdr[i]: r[i] for i in range(min(len(hdr), len(r)))})
    else:
        with open(path, newline="", encoding="utf-8-sig") as fh:
            for d in csv.DictReader(fh):
                rows.append({(k or "").strip().lower(): v for k, v in d.items()})

    out = []
    for r in rows:
        gid = str(r.get("group_id") or "").strip()
        pid = str(r.get("plan_id") or "").strip()
        mon = str(r.get("month") or "").strip()[:7]
        if not (gid and pid and mon): continue
        rec = dict(group_id=gid, plan_id=pid,
                   plan_name=str(r.get("plan_name") or "").strip(),
                   coverage=str(r.get("coverage") or "Dental").strip(),
                   month=mon,
                   submitted=_f(r.get("submitted_charges")),
                   paid=_f(r.get("paid_claims")),
                   claim_count=_i(r.get("claim_count")),
                   restated=str(r.get("restated") or "N").strip().upper().startswith("Y"),
                   enr={})
        for col, key in TIER_COLS.items():
            if r.get(col) not in (None, ""):
                rec["enr"][key] = _i(r.get(col))
        out.append(rec)
    return out


# ------------------------------------------------------------------ validate
def validate(rec, rates):
    """Return a list of problems. Empty list == row is publishable."""
    p = []
    key = f'{rec["group_id"]}/{rec["plan_id"]}'
    rt = rates.get(key)
    if rt is None:
        p.append(f"no rate table on file for {key} - cannot compute projected claims")
        return p
    if rec["submitted"] and rec["paid"] > rec["submitted"]:
        p.append(f'{rec["month"]}: paid {rec["paid"]:,.2f} exceeds submitted {rec["submitted"]:,.2f}')
    proj = sum((rec["enr"].get(t["key"]) or 0) * t["claims"] for t in rt["tiers"])
    if proj and rec["paid"] > 3*proj:
        p.append(f'{rec["month"]}: paid {rec["paid"]:,.2f} is over 3x projected '
                 f'{proj:,.2f} - confirm catch-up batch vs keying error')
    if not sum(rec["enr"].values()):
        p.append(f'{rec["month"]}: no enrollment reported')
    unknown = set(rec["enr"]) - {t["key"] for t in rt["tiers"]}
    if unknown:
        p.append(f'{rec["month"]}: enrollment in tier(s) {sorted(unknown)} not in the '
                 f'plan\'s {len(rt["tiers"])}-tier rate table')
    return p


# ------------------------------------------------------------------ merge
def merge(store, feed, rates, feed_label=None):
    """Merge feed into store. Returns (store, report)."""
    stamp = feed_label or _dt.date.today().isoformat()
    store.setdefault("months", {})
    store.setdefault("history", [])
    rep = dict(new=[], restated=[], unchanged=[], held=[])

    for rec in feed:
        problems = validate(rec, rates)
        k = f'{rec["group_id"]}|{rec["plan_id"]}|{rec["month"]}'
        if problems:
            rep["held"].append(dict(key=k, problems=problems))
            continue

        cur = store["months"].get(k)
        new = dict(enr=rec["enr"], submitted=rec["submitted"], paid=rec["paid"],
                   claim_count=rec["claim_count"], coverage=rec["coverage"],
                   plan_name=rec["plan_name"], as_of=stamp)

        if cur is None:
            store["months"][k] = new
            rep["new"].append(k)
        else:
            moved = any(abs((cur.get(f) or 0) - (new.get(f) or 0)) > 0.005
                        for f in ("submitted", "paid", "claim_count")) or \
                    cur.get("enr") != new.get("enr")
            if not moved:
                rep["unchanged"].append(k)
            else:
                store["history"].append(dict(key=k, superseded_on=stamp, was=cur))
                delta = dict(paid=round(new["paid"] - (cur.get("paid") or 0), 2),
                             submitted=round(new["submitted"] - (cur.get("submitted") or 0), 2),
                             claim_count=new["claim_count"] - (cur.get("claim_count") or 0))
                store["months"][k] = new
                rep["restated"].append(dict(key=k, delta=delta,
                                            flagged_by_sender=rec["restated"]))
    return store, rep


# ------------------------------------------------------------------ payloads
def build_payloads(store, rates, meta, plan_year_start=None, data_as_of=None):
    """Assemble generate_ppr.py payloads from stored months + CRM rates + meta.

    meta: {group_id: {client_name, plan_year_start, plan_year_label,
                      funding_basis, plans:{plan_id:{plan_name, plan_design,
                      prior_year}}}}
    """
    MONTH_ABBR = ["Jan","Feb","Mar","Apr","May","Jun",
                  "Jul","Aug","Sep","Oct","Nov","Dec"]
    by_plan = defaultdict(dict)
    for k, v in store.get("months", {}).items():
        gid, pid, mon = k.split("|")
        by_plan[(gid, pid)][mon] = v

    out = defaultdict(list)
    for (gid, pid), months in sorted(by_plan.items()):
        gm = (meta or {}).get(gid, {})
        rt = rates.get(f"{gid}/{pid}")
        if rt is None: continue
        pys = plan_year_start or gm.get("plan_year_start") or "2026-01"
        y0, m0 = (int(x) for x in pys.split("-"))
        pmeta = (gm.get("plans") or {}).get(pid, {})

        mrows, latest = [], None
        for i in range(12):
            mm = (m0 - 1 + i) % 12
            yy = y0 + (m0 - 1 + i) // 12
            key = f"{yy:04d}-{mm+1:02d}"
            src = months.get(key)
            if not src: continue
            row = dict(month_index=i, submitted=src["submitted"], paid=src["paid"],
                       claim_count=src["claim_count"])
            row.update(src["enr"])
            mrows.append(row)
            latest = f"{MONTH_ABBR[mm]} {yy}"

        if not mrows: continue
        y1 = y0 + (m0 - 1 + 11) // 12
        label = gm.get("plan_year_label") or (
            f"{MONTH_ABBR[m0-1]} {y0}\u2013{MONTH_ABBR[(m0+10)%12]} {y1}"
            if m0 != 1 else f"Jan\u2013Dec {y0}")

        out[gid].append(dict(
            group_id=gid, plan_id=pid,
            client_name=gm.get("client_name", gid),
            plan_name=pmeta.get("plan_name") or next(iter(months.values())).get("plan_name") or pid,
            coverage=next(iter(months.values())).get("coverage", "Dental"),
            plan_year_start=pys, plan_year_label=label,
            data_as_of=data_as_of or latest,
            funding_basis=gm.get("funding_basis", "claims"),
            tiers=rt["tiers"],
            plan_design=pmeta.get("plan_design", []),
            prior_year=pmeta.get("prior_year", {}),
            months=mrows))
    return dict(out)


def format_report(rep):
    L = []
    L.append(f'New months accepted:   {len(rep["new"])}')
    L.append(f'Unchanged (no-op):     {len(rep["unchanged"])}')
    L.append(f'Restated prior months: {len(rep["restated"])}')
    L.append(f'HELD for review:       {len(rep["held"])}')
    if rep["restated"]:
        L.append("\nRestatements (prior figures retained in history):")
        for r in rep["restated"]:
            d = r["delta"]
            tag = "" if r["flagged_by_sender"] else "   [NOT flagged by sender]"
            L.append(f'  {r["key"]}  paid {d["paid"]:+,.2f}  '
                     f'submitted {d["submitted"]:+,.2f}  claims {d["claim_count"]:+d}{tag}')
    if rep["held"]:
        L.append("\nHeld rows - not published:")
        for h in rep["held"]:
            for p in h["problems"]:
                L.append(f'  {h["key"]}: {p}')
    return "\n".join(L)


def main():
    if len(sys.argv) < 4:
        print(__doc__); sys.exit(1)
    feed_path, rates_path, store_path = sys.argv[1:4]
    meta_path = sys.argv[4] if len(sys.argv) > 4 else None

    rates = json.load(open(rates_path))
    store = json.load(open(store_path)) if os.path.exists(store_path) else {}
    meta  = json.load(open(meta_path)) if meta_path and os.path.exists(meta_path) else {}

    feed = read_feed(feed_path)
    print(f"Read {len(feed)} row(s) from {os.path.basename(feed_path)}\n")
    store, rep = merge(store, feed, rates)
    print(format_report(rep))

    json.dump(store, open(store_path, "w"), indent=1)
    payloads = build_payloads(store, rates, meta)
    outdir = "payloads"; os.makedirs(outdir, exist_ok=True)
    for gid, plans in payloads.items():
        fn = os.path.join(outdir, f"{gid}.json")
        json.dump(plans, open(fn, "w"), indent=1)
    print(f'\nStore updated: {store_path}')
    print(f'Payloads written for {len(payloads)} group(s) -> {outdir}/')


if __name__ == "__main__":
    main()
